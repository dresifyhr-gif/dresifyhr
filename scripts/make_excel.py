from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Travanj 2026"

COLOR_HEADER_BG = "1a1a2e"
COLOR_HEADER_FG = "e2ff54"
COLOR_ACCENT    = "e2ff54"

rows = [
    ["Dan", "Datum", "Tip objave", "Format", "Caption", "Hashtags", "Napomena", "Status"],
    [1,  "1.4.",  "Novi dres",        "Foto",            "Novi model je tu. Koji je tvoj sljedeci dres? Link u biju.",                         "#dresovi #footballdres #hrvatskinogomet #dresify",    "Objavi 10-11h",              ""],
    [2,  "2.4.",  "Reel - reveal",    "Reel",            "Ovaj je brzo otisao zadnji put... Dostupan - link u biju.",                           "#footballfashion #dresovi #reels #hrvatska",          "Brzi zoom na dres",          ""],
    [3,  "3.4.",  "Kupac recenzija",  "Foto + Story",    "Svaki tjedan novi vlasnici. Hvala nasim kupcima!",                                    "#dresify #zadovoljnikupci #dresovi",                  "Repostaj foto od kupca",     ""],
    [4,  "4.4.",  "Anketa",           "Story",           "Sto narucujes sljedece? A) Messi  B) Ronaldo",                                        "-",                                                   "Samo Story",                 ""],
    [5,  "5.4.",  "Djecji set",       "Foto",            "Tata + sin. Svejedno izgleda brutalno. Djecji set: dres + hlacice - 20 EUR.",          "#djecjidresovi #dresovi #poklonzadjecu #nogomet",    "Subota = obitelji online",   ""],
    [6,  "6.4.",  "Meme / humor",     "Foto",            "Kad prijatelji pitaju gdje si kupio dres [tag prijatelja]",                            "#footballmeme #dresovi #hrvatska",                    "Nedjelja = casual",          ""],
    [7,  "7.4.",  "Iza kulisa",       "Foto / Reel",     "Pakiranje narudzbi za danas. Ako je tvoj tu - stize ti uskoro.",                       "#dresify #narudzba #dostava",                         "Autenticno, kupci vole",     ""],
    [8,  "8.4.",  "Retro dres",       "Foto",            "Klasik koji nikad ne izlazi iz mode. Koji je tvoj najdrazi retro dres?",               "#retrodresovi #retrojersey #dresovi",                 "Retro = visok engagement",   ""],
    [9,  "9.4.",  "Vodic velicina",   "Karusel",         "Kako odabrati velicinu dresa? Vodic za odrasle i djecu.",                              "#dresovi #vodicvelicina #footballdres",               "Save & share content",       ""],
    [10, "10.4.", "Before/after",     "Reel",            "Obicna srijeda vs srijeda s novim dresom.",                                            "#dresify #footballvibes #reels",                      "Trending audio",             ""],
    [11, "11.4.", "Najava Giveaway",  "Foto + Story",    "GIVEAWAY USKORO! Pratite nas - jedan dres po vasem izboru.",                           "#giveaway #dresovi #dresify",                         "Najavi 2-3 dana prije",      ""],
    [12, "12.4.", "Giveaway objava",  "Foto",            "GIVEAWAY! 1. Prati @dresify  2. Lajkaj  3. Tag prijatelja. Dobitnik u nedjelju!",      "#giveaway #dresify #dresovi #footballgiveaway",       "Objavi ujutro",              ""],
    [13, "13.4.", "Popularni dres",   "Foto",            "Jedan od nasih najprodavanijih modela. Znate li koji je?",                             "#dresovi #bestseller #footballdres #dresify",         "Engagement pitanje",         ""],
    [14, "14.4.", "UGC recenzija",    "Foto + Story",    "Dres stigao za 2 dana, kvaliteta 10/10 - hvala!",                                     "#dresify #recenzija #zadovoljnikupci",                "Screenshot WhatsApp recenzije", ""],
    [15, "15.4.", "Giveaway rezultat","Story",           "I pobjednik je... Cestitke! Dres ide na put.",                                         "#giveaway #pobjednik #dresify",                       "Tag pobjednika",             ""],
    [16, "16.4.", "Novi model",       "Foto",            "Nije bio tu prosli tjedan. Sada jest. Link u biju.",                                   "#dresovi #novimodel #footballdres #dresify",          "",                           ""],
    [17, "17.4.", "Unboxing",         "Reel",            "POV: Tvoj dres je upravo stigao. [unboxing video]",                                    "#unboxing #dresify #dresovi #footballreels",          "Popularan format",           ""],
    [18, "18.4.", "Retro je IN",      "Karusel",         "5 razloga zasto su retro dresovi opet IN.",                                            "#retrodresovi #retrojersey #footballfashion",         "Save-worthy content",        ""],
    [19, "19.4.", "Humor",            "Reel",            "Kad izaberes velicinu a onda vidis da je ostalo samo 2 komada.",                        "#dresify #footballhumor #dresovi",                    "",                           ""],
    [20, "20.4.", "Reprezentacija",   "Foto",            "Ponos se nosi na grudima.",                                                             "#hrvatskareprezentacija #modric #dresovi #hrvatska",  "Emotivno = veci reach",      ""],
    [21, "21.4.", "Q&A",              "Story",           "Pitate - mi odgovaramo! Saljite pitanja o dresovima, dostavi, velicinama...",           "-",                                                   "Samo Story",                 ""],
    [22, "22.4.", "Nova dostava",     "Foto / Reel",     "Novi dresovi upravo stigli. Koji vam se vise svidja?",                                  "#dresify #novadostava #dresovi",                      "Curiosity hook",             ""],
    [23, "23.4.", "Testimonial",      "Foto + Story",    "Poruka koja je napravila nas dan. [screenshot recenzije]",                              "#dresify #hvala #kupci",                              "",                           ""],
    [24, "24.4.", "Trending Reel",    "Reel",            "[Trending audio] + clip dresova jedan za drugim",                                       "#reels #dresovi #footballfashion #dresify",           "Prati trending audio",       ""],
    [25, "25.4.", "Popularni klub",   "Foto",            "Nema tjedna bez jednog od ovih. Koji je tvoj izbor?",                                   "#barcelona #realmadrid #dresovi #dresify",            "Mention kluba za reach",     ""],
    [26, "26.4.", "Motivacija",       "Foto",            "Dres ne cini igraca. Ali igraca cini dres ljepsim.",                                    "#football #motivacija #dresify",                      "Light humor",                ""],
    [27, "27.4.", "Tjedni hit",       "Foto",            "Ovaj tjedan najprodavaniji je... Znate li zasto?",                                      "#dresify #bestseller #dresovi",                       "",                           ""],
    [28, "28.4.", "UGC kampanja",     "Foto + Story",    "Tagni nas u svojoj fotki s dresom! #MojDresify",                                        "#MojDresify #dresify #dresovi",                       "Pokreni branded hashtag",    ""],
    [29, "29.4.", "Humor",            "Reel",            "Prijatelji koji dijele dres vs prijatelji koji kupe svaki svoji.",                       "#footballhumor #dresify #dresovi",                    "",                           ""],
    [30, "30.4.", "Recap travnja",    "Karusel",         "Travanj u brojkama: X narudzbi, X dresova, X sretnih kupaca. Hvala vam svima!",         "#dresify #hvala #travanj #dresovi",                   "Unesi prave brojeve",        ""],
]

# Column widths
col_widths = [5, 8, 18, 14, 50, 42, 26, 13]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 30
for r in range(2, 32):
    ws.row_dimensions[r].height = 24

thin = Side(style="thin", color="d1d5db")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

FORMAT_FILL = {
    "Foto":    "dbeafe",
    "Reel":    "fce7f3",
    "Karusel": "dcfce7",
    "Story":   "fef9c3",
}

def fmt_color(fmt):
    for k, v in FORMAT_FILL.items():
        if k in fmt:
            return v
    return "f3f4f6"

# Header row
for col, val in enumerate(rows[0], 1):
    cell = ws.cell(row=1, column=col, value=val)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.font = Font(bold=True, color=COLOR_HEADER_FG, size=11, name="Calibri")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# Data rows
for row_idx, row_data in enumerate(rows[1:], 2):
    row_bg = "ffffff" if row_idx % 2 == 0 else "f9fafb"
    for col, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.border = border

        if col == 1:  # Dan
            cell.fill = PatternFill("solid", fgColor="111827")
            cell.font = Font(bold=True, color=COLOR_ACCENT, size=12, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 2:  # Datum
            cell.fill = PatternFill("solid", fgColor="1f2937")
            cell.font = Font(bold=True, color="ffffff", size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 3:  # Tip objave
            cell.fill = PatternFill("solid", fgColor="f0fdf4")
            cell.font = Font(bold=True, color="166534", size=10, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        elif col == 4:  # Format
            fc = fmt_color(str(val))
            cell.fill = PatternFill("solid", fgColor=fc)
            cell.font = Font(bold=True, color="374151", size=9, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col == 5:  # Caption
            cell.fill = PatternFill("solid", fgColor=row_bg)
            cell.font = Font(color="111827", size=9, name="Calibri")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        elif col == 6:  # Hashtags
            cell.fill = PatternFill("solid", fgColor="eff6ff")
            cell.font = Font(color="1d4ed8", size=8, name="Calibri")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        elif col == 7:  # Napomena
            cell.fill = PatternFill("solid", fgColor="fffbeb")
            cell.font = Font(italic=True, color="92400e", size=9, name="Calibri")
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        elif col == 8:  # Status
            cell.fill = PatternFill("solid", fgColor="f9fafb")
            cell.font = Font(color="374151", size=9, name="Calibri")
            cell.alignment = Alignment(horizontal="center", vertical="center")

ws.freeze_panes = "A2"

wb.save("C:/Users/vampi/Desktop/dresify-content-plan.xlsx")
print("OK")
